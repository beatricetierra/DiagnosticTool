# -*- coding: utf-8 -*-
"""
Created on Wed Dec  9 13:42:44 2020

@author: btierra
"""
import sys
import pandas as pd
import openpyxl

def NewReports(folderpath, week, *args):
    # Create empty excel files
    kvct_filepath = folderpath + 'KvctInterlocksWeek' + week + '.xlsx'
    kvct_wb = openpyxl.Workbook()
    kvct_wb.create_sheet(title='KVCT Interlocks (All)', index=0)
    kvct_wb.create_sheet(title='KVCT Interlocks (Filtered)', index=1)
    kvct_wb.save(kvct_filepath)
    
    recon_filepath = folderpath + 'ReconInterlocksWeek' + week + '.xlsx'
    recon_wb = openpyxl.Workbook()
    recon_wb.create_sheet(title='Recon Interlocks (All)', index=0)
    recon_wb.create_sheet(title='Recon Interlocks (Filtered)', index=1)
    recon_wb.save(recon_filepath)
    
    # Create empty csv file
    df = pd.DataFrame(columns=['Count', 'Interlock Number'])
    df.to_csv(folderpath+'AnalysisReportWeek' + week + '.csv', encoding='utf-8', index=False, sep='\t')
    return

def CombineExcel(combined_xlsx, append_xlsx, *args):    
    # Read Excel File
    combined = pd.ExcelFile(combined_xlsx)
    append = pd.ExcelFile(append_xlsx)
    
    # Import each sheet as a df
    combined_sheets = combined.sheet_names
    append_sheets = append.sheet_names
    
    combined_all = pd.read_excel(combined, combined_sheets[0])
    combined_filt = pd.read_excel(combined, combined_sheets[1])
    append_all = pd.read_excel(append, append_sheets[0])
    append_filt = pd.read_excel(append, append_sheets[1])
    
    # Append corresponding dataframes 
    if combined_all.empty:
        All = append_all
        Filt = append_filt
    else:
        All = combined_all.append(append_all, ignore_index=True, sort=False)
        Filt = combined_filt.append(append_filt, ignore_index=True, sort=False)
    
    try:
        All['Date'] = [date.date() for date in All['Date']]   #Remove timestamp from date format
        Filt['Date'] = [date.date() for date in Filt['Date']]
    except:
        pass
    
    # Drop duplicate rows (if function runs on same files)
    All.drop_duplicates(keep='first', inplace=True)
    Filt.drop_duplicates(keep='first', inplace=True)
    All.reset_index(inplace=True, drop=True)
    Filt.reset_index(inplace=True, drop=True)
    
    # Export and replace old kvct interlocks xlsx files
    writer = pd.ExcelWriter(combined_xlsx)
    All.to_excel(writer, sheet_name=combined_sheets[0], index=False)
    Filt.to_excel(writer, sheet_name=combined_sheets[1], index=False)
    writer.save()
    return

def CombineCSV(combined_csv, kvctpath, reconpath, *args):
    # Read filtered sheets from excel files as dataframes
    kvct = pd.ExcelFile(kvctpath)
    kvct_sheets = kvct.sheet_names
    kvct_df = pd.read_excel(kvct, kvct_sheets[1])
    
    recon = pd.ExcelFile(reconpath)
    recon_sheets = recon.sheet_names
    recon_df = pd.read_excel(recon, recon_sheets[1])
    
    # Analyze 
    dataframes = {'kvct': kvct_df, 'recon': recon_df}
    
    for key, dataframe in dataframes.items():
        analysis = dataframe.groupby('Interlock Number').count()
        analysis.reset_index(inplace=True)
        analysis = analysis[~analysis['Interlock Number'].str.contains('------')]
        analysis = analysis.iloc[:,:2]
        analysis.columns = ['Interlock Number', 'Count']
        analysis = analysis[analysis.columns[::-1]]
        dataframes[key] = analysis
    
    # Combine and export
    analysis = dataframes['kvct'].append(dataframes['recon'])
    analysis.to_csv(combined_csv, encoding='utf-8', index=False, sep='\t')
    return

if sys.argv[1] == 'create':
    NewReports(sys.argv[2], sys.argv[3])
elif sys.argv[1] == 'combinexlsx':
    CombineExcel(sys.argv[2], sys.argv[3])
elif sys.argv[1] == 'combinecsv':
    CombineCSV(sys.argv[2], sys.argv[3], sys.argv[4])

    

    
